#!/usr/bin/env python3
"""PDL toolkit: enrich/search People Data Labs Person and Company datasets.

Outputs styled .xlsx files. Use --dry-run on API subcommands to validate input
or search criteria and estimate credits without requiring PDL_API_KEY, calling
PDL, writing an xlsx, or spending credits.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

BASE = "https://api.peopledatalabs.com/v5"


def load_dotenv_value(path, key):
    path = Path(path)
    if not path.is_file():
        return None
    try:
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            if line.startswith("export "):
                line = line[len("export "):]
            k, _, v = line.partition("=")
            if k.strip() == key:
                return v.strip().strip('"').strip("'")
    except OSError:
        return None
    return None


def resolve_api_key():
    key = os.environ.get("PDL_API_KEY")
    if key:
        return key.strip()
    for candidate in (Path.cwd() / ".env", Path(__file__).resolve().parent / ".env"):
        value = load_dotenv_value(candidate, "PDL_API_KEY")
        if value:
            return value
    return None


def require_key():
    key = resolve_api_key()
    if not key:
        sys.exit(
            "No People Data Labs API key found.\n"
            "Set PDL_API_KEY as an environment variable or in a local .env file.\n"
            "Dry-run and self-test do not require a key."
        )
    return key


def pdl_request(path, api_key, params=None, json_body=None, max_retries=3):
    url = BASE + path
    for attempt in range(max_retries):
        if json_body is not None:
            data = json.dumps(json_body).encode("utf-8")
            req = urllib.request.Request(url, data=data, method="POST", headers={"X-Api-Key": api_key, "Content-Type": "application/json", "Accept": "application/json"})
        else:
            req = urllib.request.Request(url + "?" + urllib.parse.urlencode(params or {}), headers={"X-Api-Key": api_key})
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:  # noqa: S310
                return resp.status, json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            try:
                body = json.loads(e.read().decode("utf-8"))
            except Exception:
                body = {"error": {"message": str(e)}}
            if e.code in (429, 500, 502, 503) and attempt < max_retries - 1:
                time.sleep(2 ** attempt)
                continue
            return e.code, body
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)
                continue
            return 0, {"error": {"message": str(e)}}
    return 0, {"error": {"message": "exhausted retries"}}


PERSON_ALIASES = {
    "name": "name", "full name": "name", "fullname": "name",
    "first name": "first_name", "firstname": "first_name", "first": "first_name",
    "last name": "last_name", "lastname": "last_name", "last": "last_name",
    "company": "company", "organization": "company", "organisation": "company", "employer": "company", "current company": "company",
    "title": "title", "role": "title", "position": "title", "job title": "title",
    "location": "location", "city": "location", "country": "location",
    "email": "email", "e-mail": "email",
    "linkedin": "linkedin_url", "linkedin url": "linkedin_url", "profile": "linkedin_url", "linkedin profile": "linkedin_url",
}
COMPANY_ALIASES = {
    "name": "name", "company": "name", "company name": "name", "organization": "name", "organisation": "name",
    "website": "website", "domain": "website", "url": "website",
    "ticker": "ticker", "symbol": "ticker",
    "linkedin": "profile", "linkedin url": "profile", "profile": "profile",
    "location": "location", "country": "country", "region": "region", "locality": "locality", "city": "locality",
}


def load_rows(input_path, aliases):
    input_path = Path(input_path)
    suffix = input_path.suffix.lower()
    if suffix == ".txt":
        return [{"name": line.strip()} for line in input_path.read_text(encoding="utf-8").splitlines() if line.strip()]
    if suffix == ".csv":
        with input_path.open(newline="", encoding="utf-8-sig") as f:
            raw = list(csv.reader(f))
    elif suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(input_path, read_only=True, data_only=True)
        raw = [list(r) for r in wb.active.iter_rows(values_only=True)]
    else:
        raise ValueError(f"Unsupported input type: {suffix} (use .xlsx, .csv or .txt)")
    if not raw:
        return []
    headers = [aliases.get(str(h or "").strip().lower()) for h in raw[0]]
    rows = []
    for line in raw[1:]:
        row = {}
        for h, val in zip(headers, line):
            if h and val is not None and str(val).strip():
                row[h] = str(val).strip()
        if row:
            rows.append(row)
    return rows


def status_from(likelihood, min_likelihood):
    if likelihood is None:
        return "matched"
    return "needs_review" if likelihood < min_likelihood + 2 else "matched"


def contact_status(raw, values):
    if values:
        return "included"
    if raw is True:
        return "exists - upgrade to view"
    if raw is False:
        return "none on file"
    return "unknown"


def flatten_experience(experience):
    out = []
    if not isinstance(experience, list):
        return out
    for exp in experience:
        company = (exp.get("company") or {}).get("name") if isinstance(exp.get("company"), dict) else None
        title = (exp.get("title") or {}).get("name") if isinstance(exp.get("title"), dict) else None
        if company or title:
            out.append({"company": company or "", "title": title or "", "start": exp.get("start_date") or "", "end": exp.get("end_date") or ("Present" if exp.get("is_primary") else "")})
    return out


def parse_person(data, likelihood, min_likelihood, query_label, match_score=""):
    emails = []
    raw_emails = data.get("emails")
    if isinstance(raw_emails, list):
        for e in raw_emails:
            addr = e.get("address") if isinstance(e, dict) else e
            if isinstance(addr, str) and addr:
                emails.append(addr)
    work_email = data.get("work_email")
    if isinstance(work_email, str) and work_email:
        emails.insert(0, work_email)
    raw_phones = data.get("phone_numbers")
    phones = [p for p in raw_phones if isinstance(p, (str, int))] if isinstance(raw_phones, list) else []
    company = data.get("job_company_name") or ""
    if not company and isinstance(data.get("job_company"), dict):
        company = data["job_company"].get("name") or ""
    return {
        "query": query_label,
        "status": status_from(likelihood, min_likelihood),
        "match_score": match_score,
        "likelihood": likelihood if likelihood is not None else "",
        "full_name": data.get("full_name") or "",
        "job_title": data.get("job_title") or "",
        "company": company,
        "location": data.get("location_name") or "",
        "linkedin_url": data.get("linkedin_url") or "",
        "email": "; ".join(dict.fromkeys(emails)) if emails else "",
        "email_status": contact_status(raw_emails, emails),
        "phone": "; ".join(str(p) for p in phones) if phones else "",
        "phone_status": contact_status(raw_phones, phones),
        "experience": flatten_experience(data.get("experience")),
        "note": "",
    }


def empty_person(query_label, status, note=""):
    return {"query": query_label, "status": status, "match_score": "", "likelihood": "", "full_name": "", "job_title": "", "company": "", "location": "", "linkedin_url": "", "email": "", "email_status": "", "phone": "", "phone_status": "", "experience": [], "note": note}


def parse_company(data, likelihood, min_likelihood, query_label):
    loc = data.get("location") or {}
    if isinstance(loc, dict):
        hq, country = loc.get("name") or "", loc.get("country") or ""
    else:
        hq, country = (str(loc) if loc else ""), ""
    return {
        "query": query_label,
        "status": status_from(likelihood, min_likelihood),
        "likelihood": likelihood if likelihood is not None else "",
        "name": data.get("display_name") or data.get("name") or "",
        "industry": data.get("industry") or "",
        "size": data.get("size") or "",
        "employees": data.get("employee_count") or "",
        "founded": data.get("founded") or "",
        "location": hq,
        "country": country,
        "website": data.get("website") or "",
        "linkedin_url": data.get("linkedin_url") or "",
        "ticker": data.get("ticker") or "",
        "type": data.get("type") or "",
        "summary": (data.get("summary") or "")[:300],
        "note": "",
    }


def empty_company(query_label, status, note=""):
    return {"query": query_label, "status": status, "likelihood": "", "name": "", "industry": "", "size": "", "employees": "", "founded": "", "location": "", "country": "", "website": "", "linkedin_url": "", "ticker": "", "type": "", "summary": "", "note": note}


def person_label(row):
    name = row.get("name") or " ".join(filter(None, [row.get("first_name"), row.get("last_name")]))
    bits = [name]
    if row.get("company"):
        bits.append(f"@ {row['company']}")
    if not name and row.get("linkedin_url"):
        bits = [row["linkedin_url"]]
    if not name and not row.get("linkedin_url") and row.get("email"):
        bits = [row["email"]]
    return " ".join(b for b in bits if b).strip() or "(unidentified row)"


def company_label(row):
    return row.get("name") or row.get("website") or row.get("ticker") or row.get("profile") or "(unidentified row)"


STATUS_FILLS = {"matched": "E2EFDA", "needs_review": "FFF2CC", "no_match": "FCE4D6", "error": "F8CBAD"}
PEOPLE_HEADERS = ["Query", "Status", "Match score", "Likelihood", "Full name", "Current title", "Current company", "Location", "LinkedIn URL", "Email", "Email status", "Phone", "Phone status", "Most recent past role", "Note"]
PEOPLE_WIDTHS = [26, 13, 11, 11, 22, 26, 24, 22, 40, 26, 20, 16, 16, 30, 24]
COMPANY_HEADERS = ["Query", "Status", "Likelihood", "Company name", "Industry", "Size", "Employees", "Founded", "HQ location", "Country", "Website", "LinkedIn URL", "Ticker", "Type", "Summary", "Note"]
COMPANY_WIDTHS = [24, 13, 11, 26, 24, 14, 11, 9, 26, 16, 28, 38, 10, 14, 50, 24]


def _write_sheet(ws, headers, rows, status_header="Status", link_header="LinkedIn URL"):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(vertical="center")
    status_col = headers.index(status_header) + 1 if status_header in headers else None
    link_col = headers.index(link_header) + 1 if link_header in headers else None
    for row in rows:
        ws.append(row)
        ri = ws.max_row
        if status_col:
            fill = STATUS_FILLS.get(ws.cell(row=ri, column=status_col).value)
            if fill:
                ws.cell(row=ri, column=status_col).fill = PatternFill("solid", fgColor=fill)
        if link_col:
            cell = ws.cell(row=ri, column=link_col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _people_row(rec):
    most = ""
    if rec.get("experience"):
        e = rec["experience"][0]
        most = f"{e['title']} - {e['company']}".strip(" -")
    return [rec.get("query", ""), rec.get("status", ""), rec.get("match_score", ""), rec.get("likelihood", ""), rec.get("full_name", ""), rec.get("job_title", ""), rec.get("company", ""), rec.get("location", ""), rec.get("linkedin_url", ""), rec.get("email", ""), rec.get("email_status", ""), rec.get("phone", ""), rec.get("phone_status", ""), most, rec.get("note", "")]


def _company_row(rec):
    return [rec.get("query", ""), rec.get("status", ""), rec.get("likelihood", ""), rec.get("name", ""), rec.get("industry", ""), rec.get("size", ""), rec.get("employees", ""), rec.get("founded", ""), rec.get("location", ""), rec.get("country", ""), rec.get("website", ""), rec.get("linkedin_url", ""), rec.get("ticker", ""), rec.get("type", ""), rec.get("summary", ""), rec.get("note", "")]


def write_people_xlsx(records, output_path):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "People"
    _write_sheet(ws, PEOPLE_HEADERS, [_people_row(r) for r in records])
    for i, w in enumerate(PEOPLE_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    hist = wb.create_sheet("Employment history")
    hist_rows = []
    for rec in records:
        for e in rec.get("experience", []):
            hist_rows.append([rec.get("query", ""), rec.get("full_name", ""), e["company"], e["title"], e["start"], e["end"]])
    _write_sheet(hist, ["Query", "Full name", "Company", "Title", "Start", "End"], hist_rows, status_header=None, link_header=None)
    for i, w in enumerate([26, 22, 26, 26, 12, 12], 1):
        hist.column_dimensions[get_column_letter(i)].width = w
    wb.save(output_path)


def write_company_xlsx(records, output_path):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"
    _write_sheet(ws, COMPANY_HEADERS, [_company_row(r) for r in records])
    for i, w in enumerate(COMPANY_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(output_path)


def _esc(v):
    return str(v).strip().lower().replace("'", "''")


def build_person_sql(args):
    c = []
    if args.company: c.append(f"job_company_name='{_esc(args.company)}'")
    if args.title: c.append(f"job_title LIKE '%{_esc(args.title)}%'")
    if args.location: c.append(f"location_name LIKE '%{_esc(args.location)}%'")
    if args.country: c.append(f"location_country='{_esc(args.country)}'")
    if args.industry: c.append(f"industry='{_esc(args.industry)}'")
    if args.name: c.append(f"full_name LIKE '%{_esc(args.name)}%'")
    if not c:
        sys.exit("person-search needs at least one of: --company --title --location --country --industry --name")
    return "SELECT * FROM person WHERE " + " AND ".join(c)


def build_company_sql(args):
    c = []
    if args.name: c.append(f"name LIKE '%{_esc(args.name)}%'")
    if args.industry: c.append(f"industry='{_esc(args.industry)}'")
    if args.country: c.append(f"location.country='{_esc(args.country)}'")
    if args.locality: c.append(f"location.locality LIKE '%{_esc(args.locality)}%'")
    if args.tag: c.append(f"tags='{_esc(args.tag)}'")
    if args.min_employees: c.append(f"employee_count>={int(args.min_employees)}")
    if not c:
        sys.exit("company-search needs at least one of: --name --industry --country --locality --tag --min-employees")
    return "SELECT * FROM company WHERE " + " AND ".join(c)


def person_params(row, min_likelihood):
    p = {"min_likelihood": min_likelihood, "titlecase": "true"}
    for src, dst in (("linkedin_url", "profile"), ("email", "email"), ("name", "name"), ("first_name", "first_name"), ("last_name", "last_name"), ("company", "company"), ("location", "location")):
        if row.get(src): p[dst] = row[src]
    return p


def company_params(row, min_likelihood):
    p = {"min_likelihood": min_likelihood, "titlecase": "true"}
    for k in ("name", "website", "ticker", "profile", "location", "country", "region", "locality"):
        if row.get(k): p[k] = row[k]
    return p


def _dry_run_enrich(kind, rows, output):
    print("DRY RUN — no PDL API call, no credits used, no xlsx written.")
    print(f"Operation: {kind}")
    print(f"Input rows: {len(rows)}")
    print(f"Planned output: {output}")
    print(f"Estimated credits: up to {len(rows)} (PDL billing depends on endpoint terms and returned matches).")
    for row in rows[:5]:
        print("  - " + json.dumps(row, ensure_ascii=False))


def _dry_run_search(kind, sql, size, output):
    print("DRY RUN — no PDL API call, no credits used, no xlsx written.")
    print(f"Operation: {kind}")
    print(f"SQL: {sql}")
    print(f"Requested size: {size}")
    print(f"Planned output: {output}")
    print(f"Estimated credits: up to {size} (search is billed per returned record).")


def _summary(records, output_path):
    counts = {}
    for r in records:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    print(f"Done. Wrote {output_path}  ({', '.join(f'{k}: {v}' for k, v in counts.items())})")


def cmd_person_enrich(args):
    rows = load_rows(Path(args.input), PERSON_ALIASES)
    if not rows: sys.exit("No usable rows found in the input file.")
    if args.dry_run:
        _dry_run_enrich("person-enrich", rows, args.output); return
    key = require_key(); records = []
    for row in rows:
        label = person_label(row)
        sc, body = pdl_request("/person/enrich", key, params=person_params(row, args.min_likelihood))
        if sc == 200 and body.get("data"):
            records.append(parse_person(body["data"], body.get("likelihood"), args.min_likelihood, label))
        elif sc == 404:
            records.append(empty_person(label, "no_match", "no confident match in PDL"))
        else:
            records.append(empty_person(label, "error", (body.get("error") or {}).get("message") or f"HTTP {sc}"))
        time.sleep(0.2)
    write_people_xlsx(records, args.output); _summary(records, args.output)


def cmd_person_identify(args):
    rows = load_rows(Path(args.input), PERSON_ALIASES)
    if not rows: sys.exit("No usable rows found in the input file.")
    if args.dry_run:
        _dry_run_enrich("person-identify", rows, args.output); print(f"Max candidates per input row: {args.max_candidates}"); return
    key = require_key(); records = []
    for row in rows:
        label = person_label(row)
        sc, body = pdl_request("/person/identify", key, params=person_params(row, args.min_likelihood))
        matches = body.get("matches") if isinstance(body, dict) else None
        if sc == 200 and matches:
            for rank, m in enumerate(matches[:args.max_candidates], 1):
                rec = parse_person(m.get("data") or {}, None, args.min_likelihood, label, match_score=m.get("match_score", ""))
                rec["query"] = f"{label}  (candidate {rank})"; records.append(rec)
        elif sc in (200, 404):
            records.append(empty_person(label, "no_match", "no candidates returned"))
        else:
            records.append(empty_person(label, "error", (body.get("error") or {}).get("message") or f"HTTP {sc}"))
        time.sleep(0.2)
    write_people_xlsx(records, args.output); _summary(records, args.output)


def cmd_person_search(args):
    sql = args.sql or build_person_sql(args); size = max(1, min(args.size, 100))
    if args.dry_run:
        _dry_run_search("person-search", sql, size, args.output); return
    key = require_key(); body = {"sql": sql, "size": size, "dataset": args.dataset, "titlecase": True}
    sc, resp = pdl_request("/person/search", key, json_body=body)
    if sc != 200: sys.exit(f"Search failed: {(resp.get('error') or {}).get('message') or f'HTTP {sc}'}")
    label = sql.split("WHERE", 1)[-1].strip(); data = resp.get("data") or []
    records = [parse_person(d, None, args.min_likelihood, label) for d in data] or [empty_person(label, "no_match", f"0 of {resp.get('total', 0)} total matched")]
    write_people_xlsx(records, args.output); print(f"Done. {len(data)} returned of {resp.get('total', 0)} total. Wrote {args.output}")


def cmd_company_enrich(args):
    rows = load_rows(Path(args.input), COMPANY_ALIASES)
    if not rows: sys.exit("No usable rows found in the input file.")
    if args.dry_run:
        _dry_run_enrich("company-enrich", rows, args.output); return
    key = require_key(); records = []
    for row in rows:
        label = company_label(row)
        sc, body = pdl_request("/company/enrich", key, params=company_params(row, args.min_likelihood))
        if sc == 200 and (body.get("data") or body.get("name")):
            records.append(parse_company(body.get("data") or body, body.get("likelihood"), args.min_likelihood, label))
        elif sc == 404:
            records.append(empty_company(label, "no_match", "no confident match in PDL"))
        else:
            records.append(empty_company(label, "error", (body.get("error") or {}).get("message") or f"HTTP {sc}"))
        time.sleep(0.2)
    write_company_xlsx(records, args.output); _summary(records, args.output)


def cmd_company_search(args):
    sql = args.sql or build_company_sql(args); size = max(1, min(args.size, 100))
    if args.dry_run:
        _dry_run_search("company-search", sql, size, args.output); return
    key = require_key(); body = {"sql": sql, "size": size, "titlecase": True}
    sc, resp = pdl_request("/company/search", key, json_body=body)
    if sc != 200: sys.exit(f"Search failed: {(resp.get('error') or {}).get('message') or f'HTTP {sc}'}")
    label = sql.split("WHERE", 1)[-1].strip(); data = resp.get("data") or []
    records = [parse_company(d, None, args.min_likelihood, label) for d in data] or [empty_company(label, "no_match", f"0 of {resp.get('total', 0)} total matched")]
    write_company_xlsx(records, args.output); print(f"Done. {len(data)} returned of {resp.get('total', 0)} total. Wrote {args.output}")


def run_self_test(output_prefix, min_likelihood):
    person_mock = {"full_name": "Jane Tan", "job_title": "Director Of Engineering", "job_company_name": "Globex Corporation", "location_name": "Kuala Lumpur, Malaysia", "linkedin_url": "linkedin.com/in/jane-tan-example", "emails": True, "phone_numbers": False, "experience": [{"is_primary": True, "company": {"name": "Globex Corporation"}, "title": {"name": "Director of Engineering"}, "start_date": "2021-03"}, {"company": {"name": "Initech"}, "title": {"name": "Engineering Manager"}, "start_date": "2017-01", "end_date": "2021-02"}]}
    company_mock = {"display_name": "Globex Corporation", "industry": "information technology and services", "size": "1001-5000", "employee_count": 3710, "founded": 1999, "location": {"name": "kuala lumpur, malaysia", "country": "malaysia"}, "website": "globex.com", "linkedin_url": "linkedin.com/company/globex", "ticker": "GLBX", "type": "public", "summary": "A diversified technology company."}
    write_people_xlsx([parse_person(person_mock, 8, min_likelihood, "Jane Tan @ Globex"), empty_person("Nonexistent Person @ Nowhere", "no_match", "PDL 404")], f"{output_prefix}_people.xlsx")
    write_company_xlsx([parse_company(company_mock, 9, min_likelihood, "Globex"), empty_company("Nowhere Inc", "no_match", "PDL 404")], f"{output_prefix}_companies.xlsx")
    print(f"[self-test] wrote {output_prefix}_people.xlsx and {output_prefix}_companies.xlsx")


KNOWN = {"person-enrich", "person-identify", "person-search", "company-enrich", "company-search"}


def build_parser():
    p = argparse.ArgumentParser(description="People Data Labs enrichment & search toolkit.")
    p.add_argument("--self-test", action="store_true", help="Offline demo (no key/network).")
    sub = p.add_subparsers(dest="command")
    def add_common(sp, default_out):
        sp.add_argument("--output", default=default_out)
        sp.add_argument("--min-likelihood", type=int, default=4)
        sp.add_argument("--dry-run", action="store_true", help="Validate input/search and estimate credits without calling PDL or requiring PDL_API_KEY.")
    pe = sub.add_parser("person-enrich"); pe.add_argument("--input", required=True); add_common(pe, "enriched_people.xlsx"); pe.set_defaults(func=cmd_person_enrich)
    pi = sub.add_parser("person-identify"); pi.add_argument("--input", required=True); pi.add_argument("--max-candidates", type=int, default=5); add_common(pi, "person_candidates.xlsx"); pi.set_defaults(func=cmd_person_identify)
    ps = sub.add_parser("person-search")
    for f in ("company", "title", "location", "country", "industry", "name"): ps.add_argument(f"--{f}")
    ps.add_argument("--size", type=int, default=10); ps.add_argument("--dataset", default="all"); ps.add_argument("--sql"); add_common(ps, "person_search.xlsx"); ps.set_defaults(func=cmd_person_search)
    ce = sub.add_parser("company-enrich"); ce.add_argument("--input", required=True); add_common(ce, "enriched_companies.xlsx"); ce.set_defaults(func=cmd_company_enrich)
    cs = sub.add_parser("company-search")
    for f in ("name", "industry", "country", "locality", "tag"): cs.add_argument(f"--{f}")
    cs.add_argument("--min-employees", type=int); cs.add_argument("--size", type=int, default=10); cs.add_argument("--sql"); add_common(cs, "company_search.xlsx"); cs.set_defaults(func=cmd_company_search)
    return p


def main():
    for stream in (sys.stdout, sys.stderr):
        try: stream.reconfigure(encoding="utf-8")
        except (AttributeError, ValueError): pass
    argv = sys.argv[1:]; first = argv[0].lower() if argv else ""
    if argv and first not in KNOWN and not first.startswith("person") and not first.startswith("company"):
        if "--self-test" in argv:
            tmp = argparse.ArgumentParser(add_help=False); tmp.add_argument("--self-test", action="store_true"); tmp.add_argument("--output", default="selftest"); tmp.add_argument("--min-likelihood", type=int, default=4)
            ns, _ = tmp.parse_known_args(argv); run_self_test(ns.output if ns.output != "selftest" else "selftest", ns.min_likelihood); return
        if "--input" in argv: argv = ["person-enrich"] + argv
    parser = build_parser(); args = parser.parse_args(argv)
    if getattr(args, "self_test", False): run_self_test("selftest", getattr(args, "min_likelihood", 4)); return
    if not getattr(args, "command", None): parser.print_help(); return
    args.func(args)


if __name__ == "__main__":
    main()
