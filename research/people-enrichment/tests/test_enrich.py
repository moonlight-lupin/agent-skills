#!/usr/bin/env python3
"""Unit tests for enrich.py — PDL enrichment & search toolkit.

Run:  python -m pytest research/people-enrichment/tests/test_enrich.py -v
  or:  python research/people-enrichment/tests/test_enrich.py

No network, no API key required — all tests use synthetic fixtures and mock data.
"""

import sys
import csv
import tempfile
import argparse
import unittest
from pathlib import Path

SCRIPTS_DIR = Path(__file__).resolve().parent.parent / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

import enrich  # noqa: E402
import openpyxl  # noqa: E402


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_csv(path, headers, rows):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(str(path), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for row in rows:
            w.writerow(row)


def _make_xlsx(path, headers, rows):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(str(path))


def _make_txt(path, lines):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_text("\n".join(lines) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# status_from — confidence threshold logic
# ---------------------------------------------------------------------------

class TestStatusFrom(unittest.TestCase):
    """status_from: likelihood → matched/needs_review based on min_likelihood."""

    def test_none_likelihood_is_matched(self):
        self.assertEqual(enrich.status_from(None, 4), "matched")

    def test_high_likelihood_matched(self):
        self.assertEqual(enrich.status_from(8, 4), "matched")  # 8 >= 4+2

    def test_at_boundary_matched(self):
        """likelihood == min + 2 → matched (boundary is >=)."""
        self.assertEqual(enrich.status_from(6, 4), "matched")  # 6 >= 6

    def test_below_boundary_needs_review(self):
        self.assertEqual(enrich.status_from(5, 4), "needs_review")  # 5 < 6

    def test_low_likelihood_needs_review(self):
        self.assertEqual(enrich.status_from(3, 4), "needs_review")

    def test_high_min_likelihood(self):
        """With min=8, likelihood=9 → needs_review (9 < 10)."""
        self.assertEqual(enrich.status_from(9, 8), "needs_review")
        self.assertEqual(enrich.status_from(10, 8), "matched")


# ---------------------------------------------------------------------------
# contact_status — free-plan PII boolean logic
# ---------------------------------------------------------------------------

class TestContactStatus(unittest.TestCase):
    """contact_status: 4 PII states (included, exists-upgrade, none, unknown)."""

    def test_included(self):
        """Real values present (Pro plan)."""
        self.assertEqual(enrich.contact_status(["a@b.com"], ["a@b.com"]), "included")

    def test_exists_upgrade_to_view(self):
        """Free plan: True = exists but paywalled."""
        self.assertEqual(enrich.contact_status(True, []), "exists - upgrade to view")

    def test_none_on_file(self):
        """Free plan: False = no contact on file."""
        self.assertEqual(enrich.contact_status(False, []), "none on file")

    def test_unknown(self):
        """No raw value at all."""
        self.assertEqual(enrich.contact_status(None, []), "unknown")

    def test_empty_list_raw_is_unknown(self):
        """Empty list raw → unknown (not 'none on file')."""
        self.assertEqual(enrich.contact_status([], []), "unknown")

    def test_values_present_regardless_of_raw(self):
        """If values list is non-empty, status is 'included' even if raw is True."""
        self.assertEqual(enrich.contact_status(True, ["a@b.com"]), "included")


# ---------------------------------------------------------------------------
# flatten_experience — nested experience extraction
# ---------------------------------------------------------------------------

class TestFlattenExperience(unittest.TestCase):
    """flatten_experience: nested dict → flat list, is_primary → Present."""

    def test_basic(self):
        exp = [
            {"is_primary": True, "company": {"name": "Globex"},
             "title": {"name": "Director"}, "start_date": "2021-03"},
            {"company": {"name": "Initech"}, "title": {"name": "Manager"},
             "start_date": "2017-01", "end_date": "2021-02"},
        ]
        result = enrich.flatten_experience(exp)
        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]["company"], "Globex")
        self.assertEqual(result[0]["title"], "Director")
        self.assertEqual(result[0]["end"], "Present")  # is_primary → Present
        self.assertEqual(result[1]["end"], "2021-02")

    def test_non_list_returns_empty(self):
        self.assertEqual(enrich.flatten_experience(None), [])
        self.assertEqual(enrich.flatten_experience("not a list"), [])

    def test_empty_list(self):
        self.assertEqual(enrich.flatten_experience([]), [])

    def test_missing_company_or_title_skipped(self):
        """Entry with neither company nor title → skipped."""
        exp = [{"start_date": "2020-01"}]  # no company, no title
        self.assertEqual(enrich.flatten_experience(exp), [])

    def test_partial_company(self):
        """Company present, title missing → included with empty title."""
        exp = [{"company": {"name": "Globex"}, "start_date": "2021-03"}]
        result = enrich.flatten_experience(exp)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["company"], "Globex")
        self.assertEqual(result[0]["title"], "")

    def test_company_as_string(self):
        """Company field as a string (not dict) → handled."""
        exp = [{"company": "Globex", "title": {"name": "Director"}, "start_date": "2021-03"}]
        result = enrich.flatten_experience(exp)
        # When company is not a dict, company_name = None → entry has company="" but title="Director"
        self.assertEqual(len(result), 1)


# ---------------------------------------------------------------------------
# parse_person — full person record assembly
# ---------------------------------------------------------------------------

class TestParsePerson(unittest.TestCase):
    """parse_person: email dedup, work_email priority, boolean PII, experience."""

    def test_pro_plan_emails(self):
        """Pro plan: emails as list of dicts with addresses."""
        data = {
            "full_name": "Jane Tan", "job_title": "Director",
            "job_company_name": "Globex", "location_name": "KL",
            "emails": [{"address": "jane@globex.com"}, {"address": "jane@example.com"}],
            "phone_numbers": ["+601****5678"],
            "experience": [],
        }
        rec = enrich.parse_person(data, 8, 4, "Jane Tan @ Globex")
        self.assertEqual(rec["status"], "matched")
        self.assertIn("jane@globex.com", rec["email"])
        self.assertEqual(rec["email_status"], "included")
        self.assertIn("+601****5678", rec["phone"])
        self.assertEqual(rec["phone_status"], "included")

    def test_work_email_priority(self):
        """work_email is inserted at position 0 in the email list."""
        data = {
            "full_name": "Jane", "emails": [{"address": "a@b.com"}],
            "work_email": "work@company.com",
        }
        rec = enrich.parse_person(data, 8, 4, "Jane")
        emails = rec["email"].split("; ")
        self.assertEqual(emails[0], "work@company.com")

    def test_email_dedup(self):
        """Duplicate emails are deduplicated via dict.fromkeys."""
        data = {
            "full_name": "Jane",
            "emails": [{"address": "dup@example.com"}, {"address": "dup@example.com"}],
        }
        rec = enrich.parse_person(data, 8, 4, "Jane")
        self.assertEqual(rec["email"].count("dup@example.com"), 1)

    def test_free_plan_boolean_emails(self):
        """Free plan: emails=True (paywalled) → status 'exists - upgrade to view'."""
        data = {"full_name": "Jane", "emails": True, "phone_numbers": False}
        rec = enrich.parse_person(data, 5, 4, "Jane")
        self.assertEqual(rec["email"], "")
        self.assertEqual(rec["email_status"], "exists - upgrade to view")
        self.assertEqual(rec["phone"], "")
        self.assertEqual(rec["phone_status"], "none on file")

    def test_emails_as_strings(self):
        """emails list of strings (not dicts) → handled."""
        data = {"full_name": "Jane", "emails": ["a@b.com", "c@d.com"]}
        rec = enrich.parse_person(data, 8, 4, "Jane")
        self.assertIn("a@b.com", rec["email"])
        self.assertIn("c@d.com", rec["email"])

    def test_no_experience(self):
        """Missing or None experience → empty list."""
        rec = enrich.parse_person({"full_name": "Jane", "experience": None}, 8, 4, "Jane")
        self.assertEqual(rec["experience"], [])

    def test_empty_experience_list(self):
        rec = enrich.parse_person({"full_name": "Jane", "experience": []}, 8, 4, "Jane")
        self.assertEqual(rec["experience"], [])

    def test_job_company_fallback(self):
        """job_company dict fallback when job_company_name is absent."""
        data = {"full_name": "Jane", "job_company": {"name": "Globex"}}
        rec = enrich.parse_person(data, 8, 4, "Jane")
        self.assertEqual(rec["company"], "Globex")

    def test_match_score_passthrough(self):
        rec = enrich.parse_person({"full_name": "Jane"}, 8, 4, "Jane", match_score="0.95")
        self.assertEqual(rec["match_score"], "0.95")


# ---------------------------------------------------------------------------
# parse_company — company record assembly
# ---------------------------------------------------------------------------

class TestParseCompany(unittest.TestCase):
    """parse_company: location dict/string, summary truncation, name fallback."""

    def test_basic(self):
        data = {
            "display_name": "Globex Corp", "industry": "tech", "size": "1001-5000",
            "employee_count": 3710, "founded": 1999,
            "location": {"name": "KL, Malaysia", "country": "Malaysia"},
            "website": "globex.com", "linkedin_url": "linkedin.com/company/globex",
            "ticker": "GLBX", "type": "public", "summary": "A tech company.",
        }
        rec = enrich.parse_company(data, 9, 4, "Globex")
        self.assertEqual(rec["name"], "Globex Corp")
        self.assertEqual(rec["location"], "KL, Malaysia")
        self.assertEqual(rec["country"], "Malaysia")
        self.assertEqual(rec["status"], "matched")

    def test_location_as_string(self):
        data = {"display_name": "Test", "location": "Kuala Lumpur, Malaysia"}
        rec = enrich.parse_company(data, 8, 4, "Test")
        self.assertEqual(rec["location"], "Kuala Lumpur, Malaysia")
        self.assertEqual(rec["country"], "")

    def test_location_none(self):
        data = {"display_name": "Test", "location": None}
        rec = enrich.parse_company(data, 8, 4, "Test")
        self.assertEqual(rec["location"], "")
        self.assertEqual(rec["country"], "")

    def test_summary_truncation(self):
        data = {"display_name": "Test", "summary": "A" * 500}
        rec = enrich.parse_company(data, 8, 4, "Test")
        self.assertEqual(len(rec["summary"]), 300)

    def test_display_name_over_name(self):
        """display_name takes priority over name."""
        data = {"name": "Primary", "display_name": "Display"}
        rec = enrich.parse_company(data, 8, 4, "Test")
        self.assertEqual(rec["name"], "Display")

    def test_name_fallback(self):
        """Falls back to 'name' when display_name is absent."""
        data = {"name": "Primary"}
        rec = enrich.parse_company(data, 8, 4, "Test")
        self.assertEqual(rec["name"], "Primary")


# ---------------------------------------------------------------------------
# person_label / company_label — query label construction
# ---------------------------------------------------------------------------

class TestPersonLabel(unittest.TestCase):
    """person_label: name + @ company, fallbacks for missing name."""

    def test_name_and_company(self):
        self.assertEqual(enrich.person_label({"name": "Jane", "company": "Globex"}), "Jane @ Globex")

    def test_first_and_last(self):
        self.assertEqual(enrich.person_label({"first_name": "Jane", "last_name": "Tan"}), "Jane Tan")

    def test_linkedin_only(self):
        self.assertEqual(enrich.person_label({"linkedin_url": "linkedin.com/in/jane"}), "linkedin.com/in/jane")

    def test_email_only(self):
        self.assertEqual(enrich.person_label({"email": "jane@b.com"}), "jane@b.com")

    def test_empty(self):
        self.assertEqual(enrich.person_label({}), "(unidentified row)")


class TestCompanyLabel(unittest.TestCase):
    """company_label: name → website → ticker → profile → fallback."""

    def test_name(self):
        self.assertEqual(enrich.company_label({"name": "Globex"}), "Globex")

    def test_website_fallback(self):
        self.assertEqual(enrich.company_label({"website": "globex.com"}), "globex.com")

    def test_ticker_fallback(self):
        self.assertEqual(enrich.company_label({"ticker": "GLBX"}), "GLBX")

    def test_profile_fallback(self):
        self.assertEqual(enrich.company_label({"profile": "linkedin.com/company/globex"}),
                         "linkedin.com/company/globex")

    def test_empty(self):
        self.assertEqual(enrich.company_label({}), "(unidentified row)")


# ---------------------------------------------------------------------------
# _esc — SQL escaping
# ---------------------------------------------------------------------------

class TestEsc(unittest.TestCase):
    """_esc: lowercase, strip, double single-quotes."""

    def test_lowercases(self):
        self.assertEqual(enrich._esc("HELLO"), "hello")

    def test_strips(self):
        self.assertEqual(enrich._esc("  Hello  "), "hello")

    def test_doubles_single_quotes(self):
        self.assertEqual(enrich._esc("O'Brien"), "o''brien")

    def test_non_string(self):
        self.assertEqual(enrich._esc(42), "42")


# ---------------------------------------------------------------------------
# build_person_sql / build_company_sql — flag → SQL
# ---------------------------------------------------------------------------

class TestBuildPersonSql(unittest.TestCase):
    """build_person_sql: flags combined with AND, LIKE for fuzzy fields."""

    def test_company_exact(self):
        args = argparse.Namespace(company="Northwind", title=None, location=None,
                                   country=None, industry=None, name=None)
        sql = enrich.build_person_sql(args)
        self.assertIn("job_company_name='northwind'", sql)

    def test_title_like(self):
        args = argparse.Namespace(company=None, title="director", location=None,
                                   country=None, industry=None, name=None)
        sql = enrich.build_person_sql(args)
        self.assertIn("job_title LIKE '%director%'", sql)

    def test_multiple_flags_and(self):
        args = argparse.Namespace(company="Globex", title="director", location="KL",
                                   country=None, industry=None, name=None)
        sql = enrich.build_person_sql(args)
        self.assertIn(" AND ", sql)
        self.assertIn("job_company_name='globex'", sql)
        self.assertIn("job_title LIKE '%director%'", sql)
        self.assertIn("location_name LIKE '%kl%'", sql)

    def test_no_flags_exits(self):
        args = argparse.Namespace(company=None, title=None, location=None,
                                   country=None, industry=None, name=None)
        with self.assertRaises(SystemExit):
            enrich.build_person_sql(args)


class TestBuildCompanySql(unittest.TestCase):
    """build_company_sql: flags → SQL, min_employees as int."""

    def test_min_employees(self):
        args = argparse.Namespace(name=None, industry=None, country=None, locality=None,
                                   tag=None, min_employees=50)
        sql = enrich.build_company_sql(args)
        self.assertIn("employee_count>=50", sql)

    def test_no_flags_exits(self):
        args = argparse.Namespace(name=None, industry=None, country=None, locality=None,
                                   tag=None, min_employees=None)
        with self.assertRaises(SystemExit):
            enrich.build_company_sql(args)


# ---------------------------------------------------------------------------
# load_rows — input parsing
# ---------------------------------------------------------------------------

class TestLoadRows(unittest.TestCase):
    """load_rows: CSV, XLSX, TXT, aliases, empty rows."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def test_csv(self):
        path = Path(self.tmp) / "people.csv"
        _make_csv(path, ["name", "company", "email"], [["Jane", "Globex", "j@g.com"], ["Arjun", "Initech", ""]])
        rows = enrich.load_rows(path, enrich.PERSON_ALIASES)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["name"], "Jane")
        self.assertEqual(rows[0]["company"], "Globex")
        # Empty cell should be omitted
        self.assertNotIn("email", rows[1])

    def test_csv_alias_columns(self):
        path = Path(self.tmp) / "aliases.csv"
        _make_csv(path, ["Full Name", "Employer", "Job Title"], [["Jane Tan", "Globex", "Director"]])
        rows = enrich.load_rows(path, enrich.PERSON_ALIASES)
        self.assertEqual(rows[0]["name"], "Jane Tan")
        self.assertEqual(rows[0]["company"], "Globex")
        self.assertEqual(rows[0]["title"], "Director")

    def test_xlsx(self):
        path = Path(self.tmp) / "people.xlsx"
        _make_xlsx(path, ["Name", "Company"], [["Jane", "Globex"], ["Arjun", "Initech"]])
        rows = enrich.load_rows(path, enrich.PERSON_ALIASES)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["name"], "Jane")

    def test_txt_one_name_per_line(self):
        path = Path(self.tmp) / "names.txt"
        _make_txt(path, ["Jane Tan", "Arjun Mehta", "", "  "])
        rows = enrich.load_rows(path, enrich.PERSON_ALIASES)
        self.assertEqual(len(rows), 2)  # blank/whitespace lines skipped
        self.assertEqual(rows[0]["name"], "Jane Tan")

    def test_unsupported_type(self):
        path = Path(self.tmp) / "data.json"
        path.write_text("{}")
        with self.assertRaises(ValueError):
            enrich.load_rows(path, enrich.PERSON_ALIASES)

    def test_empty_file(self):
        path = Path(self.tmp) / "empty.csv"
        _make_csv(path, ["name"], [])
        rows = enrich.load_rows(path, enrich.PERSON_ALIASES)
        self.assertEqual(rows, [])

    def test_company_aliases(self):
        path = Path(self.tmp) / "companies.csv"
        _make_csv(path, ["Company Name", "Website"], [["Globex", "globex.com"]])
        rows = enrich.load_rows(path, enrich.COMPANY_ALIASES)
        self.assertEqual(rows[0]["name"], "Globex")
        self.assertEqual(rows[0]["website"], "globex.com")


# ---------------------------------------------------------------------------
# load_dotenv_value — .env parsing
# ---------------------------------------------------------------------------

class TestLoadDotenv(unittest.TestCase):
    """load_dotenv_value: basic parsing, export prefix, quotes, comments."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def test_basic(self):
        path = Path(self.tmp) / ".env"
        path.write_text("PDL_API_KEY=abc123\n")
        self.assertEqual(enrich.load_dotenv_value(path, "PDL_API_KEY"), "abc123")

    def test_export_prefix(self):
        """'export KEY=VALUE' is handled (the #3 fix)."""
        path = Path(self.tmp) / ".env"
        path.write_text("export PDL_API_KEY=abc123\n")
        self.assertEqual(enrich.load_dotenv_value(path, "PDL_API_KEY"), "abc123")

    def test_quotes_stripped(self):
        path = Path(self.tmp) / ".env"
        path.write_text('PDL_API_KEY="abc123"\n')
        self.assertEqual(enrich.load_dotenv_value(path, "PDL_API_KEY"), "abc123")

    def test_single_quotes_stripped(self):
        path = Path(self.tmp) / ".env"
        path.write_text("PDL_API_KEY='abc123'\n")
        self.assertEqual(enrich.load_dotenv_value(path, "PDL_API_KEY"), "abc123")

    def test_comment_ignored(self):
        path = Path(self.tmp) / ".env"
        path.write_text("# comment\nPDL_API_KEY=abc123\n")
        self.assertEqual(enrich.load_dotenv_value(path, "PDL_API_KEY"), "abc123")

    def test_no_file_returns_none(self):
        self.assertIsNone(enrich.load_dotenv_value(Path(self.tmp) / "nope.env", "KEY"))

    def test_wrong_key_returns_none(self):
        path = Path(self.tmp) / ".env"
        path.write_text("OTHER_KEY=abc\n")
        self.assertIsNone(enrich.load_dotenv_value(path, "PDL_API_KEY"))


# ---------------------------------------------------------------------------
# empty_person / empty_company
# ---------------------------------------------------------------------------

class TestEmptyRecords(unittest.TestCase):
    """empty_person / empty_company: status + note, all fields blank."""

    def test_empty_person(self):
        rec = enrich.empty_person("Test", "no_match", "not found")
        self.assertEqual(rec["status"], "no_match")
        self.assertEqual(rec["note"], "not found")
        self.assertEqual(rec["full_name"], "")
        self.assertEqual(rec["email"], "")

    def test_empty_company(self):
        rec = enrich.empty_company("Test", "error", "no data")
        self.assertEqual(rec["status"], "error")
        self.assertEqual(rec["note"], "no data")
        self.assertEqual(rec["name"], "")


# ---------------------------------------------------------------------------
# KNOWN subcommands + backward compat
# ---------------------------------------------------------------------------

class TestKnownSubcommands(unittest.TestCase):
    """KNOWN set and subcommand routing."""

    def test_five_commands(self):
        self.assertEqual(len(enrich.KNOWN), 5)
        for cmd in ("person-enrich", "person-identify", "person-search",
                    "company-enrich", "company-search"):
            self.assertIn(cmd, enrich.KNOWN)


# ---------------------------------------------------------------------------
# xlsx writing — structural verification
# ---------------------------------------------------------------------------

class TestXlsxWriting(unittest.TestCase):
    """write_people_xlsx / write_company_xlsx: headers, tabs, data rows."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def test_people_xlsx_two_tabs(self):
        rec = enrich.parse_person(
            {"full_name": "Jane", "job_title": "Director", "job_company_name": "Globex",
             "emails": [{"address": "j@g.com"}],
             "experience": [{"is_primary": True, "company": {"name": "Globex"},
                             "title": {"name": "Director"}, "start_date": "2021-03"}]},
            8, 4, "Jane @ Globex")
        path = str(Path(self.tmp) / "people.xlsx")
        enrich.write_people_xlsx([rec], path)
        wb = openpyxl.load_workbook(path)
        self.assertIn("People", wb.sheetnames)
        self.assertIn("Employment history", wb.sheetnames)

    def test_company_xlsx_single_tab(self):
        rec = enrich.parse_company(
            {"display_name": "Globex", "industry": "tech", "size": "100",
             "employee_count": 100, "founded": 2000,
             "location": {"name": "KL", "country": "Malaysia"},
             "website": "globex.com"}, 8, 4, "Globex")
        path = str(Path(self.tmp) / "companies.xlsx")
        enrich.write_company_xlsx([rec], path)
        wb = openpyxl.load_workbook(path)
        self.assertEqual(len(wb.sheetnames), 1)
        self.assertEqual(wb.sheetnames[0], "Companies")


# ---------------------------------------------------------------------------
# Self-test (offline)
# ---------------------------------------------------------------------------

class TestSelfTest(unittest.TestCase):
    """run_self_test produces valid xlsx files."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def test_writes_two_files(self):
        prefix = str(Path(self.tmp) / "st")
        enrich.run_self_test(prefix, 4)
        people_path = Path(f"{prefix}_people.xlsx")
        companies_path = Path(f"{prefix}_companies.xlsx")
        self.assertTrue(people_path.is_file())
        self.assertTrue(companies_path.is_file())
        # Verify people xlsx has correct headers
        wb = openpyxl.load_workbook(str(people_path))
        ws = wb["People"]
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers, enrich.PEOPLE_HEADERS)


if __name__ == "__main__":
    unittest.main(verbosity=2)